from flask import Flask, jsonify, request, render_template
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
import os
from dotenv import load_dotenv

load_dotenv()


app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
ma = Marshmallow(app)


class Formulas(db.Model):
    __tablename__ = "formulas"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    formula = db.Column(db.String(255), nullable=False)

    def __repr__(self):
        return f"<User {self.name}>,{self.formula}"


class FormulaSchema(ma.Schema):
    class Meta:
        fields = ["__all__"]


formula_schemas = FormulaSchema(many=True)


@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")


@app.route("/show", methods=["GET"])
def show_users():
    forms = Formulas.query.all()
    return render_template("view.html", forms=forms)


@app.route("/add", methods=["POST"])
def add_data():
    if request.method == "POST":
        exc_data = request.files["formula"]

        if not exc_data.filename.endswith(".xlsx"):
            return "Error: Please upload an Excel file."

        try:
            print(exc_data)

            Mydata = load_workbook(exc_data)
            Newdata = Mydata.active

            for row in Newdata.iter_rows(min_row=2, values_only=True):
                # print("Row Data:", row)
                data = Formulas(id=row[0], name=row[1], formula=row[2])
                db.session.add(data)

            db.session.commit()

        except Exception as e:
            return f"Error occurred while processing the file: {str(e)}"

    return "message : Data retrieved"


with app.app_context():
    db.create_all()
if __name__ == "__main__":
    app.run(debug=True)
